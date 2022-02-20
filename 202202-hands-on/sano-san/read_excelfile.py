# PythonでExcelファイルを読み込んで、CSVファイルに書き出してみる
import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

HANDSON_DIR = Path(__file__).parent
EXAMPLE_EXCEL_FILEPATH = HANDSON_DIR / "./example_tyohyo.xlsx"
# print(EXAMPLE_EXCEL_FILEPATH)

EXPORT_JSON_FILEPATH = HANDSON_DIR / "export_tyohyo.json"


def main() -> None:
    # （Excelの数式はすでに計算された結果が入っていることを想定）
    ex_wb = load_workbook(filename=EXAMPLE_EXCEL_FILEPATH, data_only=True)

    print(f"ワークシート一覧: {ex_wb.sheetnames}")

    # シート一つなので activeで有効にします。
    ex_ws = ex_wb.active
    ex_ws_value = ex_ws["A1"].value

    print(f"セルA1の値: {ex_ws_value}")

    # 読み込みたいセル番号を指定してみましょう
    celladdr_employee_number = "B3"
    celladdr_employee_name = "B4"
    celladdr_statement_number = "E3"
    celladdr_application_day = "E4"
    celladdr_total_amount = "D6"

    cell_employee_number = ex_ws[celladdr_employee_number]
    cell_employee_name = ex_ws[celladdr_employee_name]
    cell_statement_number = ex_ws[celladdr_statement_number]
    # 注意: Excelで日付として表現されている値はdatetimeオブジェクトに変換されます
    cell_application_day = ex_ws[celladdr_application_day]
    cell_total_amount = ex_ws[celladdr_total_amount]

    print(
        f"""従業員番号:{cell_employee_number.value}
申請者:{cell_employee_name.value}
明細書番号:{cell_statement_number.value}
申請日:{cell_application_day.value.isoformat(" ")}
合計金額:{cell_total_amount.value}
    """
    )

    # セル範囲を指定して読み込んでみましょう

    cell_koumoku_list = ex_ws["A9":"D23"]
    from pprint import pprint
    pprint(cell_koumoku_list)
    column_names = [
        "日付",
        "内容",
        "支払先",
        "金額",
        "備考",
    ]

    for row in cell_koumoku_list:
        for column_name, column_cell in zip(column_names, row):
            if column_cell.value is not None:
                print(f"{column_name}:{column_cell.value},")
        # print("\n")

    # TODO:2022-02-19 2022-02はここまででした。

#     # jsonファイルとして書き出してみましょう
#     # jsonデータの構造を用意
#     export_json_str = """
# {
#     "employee_number": "",
#     "employee_name": "",
#     "statement_number": "",
#     "application_day": "",
#     "total_amount": "",
#     "koumoku_list": []
# }
#     """
#     export_json_data = json.loads(export_json_str)

#     # 明細書自体の情報を追加
#     export_json_data["employee_number"] = cell_employee_number.value
#     export_json_data["employee_name"] = cell_employee_name.value
#     export_json_data["statement_number"] = cell_statement_number.value

#     # datetimeオブジェクトを文字列化
#     export_cell_application_day_str = cell_application_day.value.isoformat(" ")
#     export_json_data["application_day"] = export_cell_application_day_str
#     export_json_data["total_amount"] = cell_total_amount.value

#     # 明細書項目一覧を生成
#     for row in cell_koumoku_list:
#         # columnの辞書の初期化
#         json_row_data = {}
#         for column_name, column_cell in zip(column_names, row):
#             # jsonのrowへ値を追加する
#             if column_cell.value is not None:
#                 # columnの辞書にrowの値を追加
#                 column_cell_value = column_cell.value
#                 # datetimeがある場合の処理
#                 if isinstance(column_cell_value, datetime):
#                     column_cell_value = column_cell_value.isoformat(" ")
#                 json_row_data |= {column_name: column_cell_value}
#         # 要素が全くない場合
#         if json_row_data:
#             export_json_data["koumoku_list"].append(json_row_data)

#     # 表示する
#     from pprint import pprint

#     pprint(export_json_data)

#     # jsonファイルで保存しましょう
#     with EXPORT_JSON_FILEPATH.open("w") as export_json_file:
#         json.dump(
#             export_json_data,
#             export_json_file,
#             indent=4,
#             ensure_ascii=False,
#         )


if __name__ == "__main__":
    main()
